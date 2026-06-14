import openpyxl

for fname in ['K_TAHP_Result.xlsx', 'K_FAHP_Result.xlsx']:
    print(f"\n{'='*60}")
    print(f"FILE: {fname}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(fname)
    print(f"Sheets: {wb.sheetnames}")
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n--- Sheet: {sname} (rows={ws.max_row}, cols={ws.max_column}) ---")
        for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True):
            print(list(row))
