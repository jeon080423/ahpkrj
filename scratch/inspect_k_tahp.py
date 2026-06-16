import openpyxl

wb = openpyxl.load_workbook("G:\\K_TAHP_Result.xlsx")
ws = wb["종합분석"]

print("Row 4 (Headers):")
row_vals = [ws.cell(row=4, column=c).value for c in range(1, 10)]
print(row_vals)

print("Row 9:")
row_vals = [ws.cell(row=9, column=c).value for c in range(1, 10)]
print(row_vals)

print("Row 10:")
row_vals = [ws.cell(row=10, column=c).value for c in range(1, 10)]
print(row_vals)
